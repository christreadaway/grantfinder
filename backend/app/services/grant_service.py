import io
from typing import List, Optional
from datetime import datetime
from openpyxl import load_workbook


class GrantService:
    """Service for parsing and managing grant databases."""

    # Expected column mappings (flexible naming)
    COLUMN_MAPPINGS = {
        "name": ["grant name", "name", "grant", "title"],
        "granting_authority": ["granting authority", "authority", "grantor", "foundation", "organization"],
        "deadline": ["deadline", "due date", "due", "application deadline"],
        "amount_min": ["amount min", "min amount", "minimum", "min"],
        "amount_max": ["amount max", "max amount", "maximum", "max", "up to", "amount"],
        "description": ["description", "purpose", "desc", "what it funds", "overview"],
        "eligibility": ["eligibility", "requirements", "who can apply", "eligible"],
        "apply_url": ["apply url", "url", "link", "application link", "website", "apply"],
        "geographic_restriction": ["geographic restriction", "geography", "state", "region", "location"],
        "categories": ["categories", "category", "tags", "type", "grant type"],
        "notes": ["notes", "note", "additional info", "comments"],
    }

    @staticmethod
    async def parse_excel(file_content: bytes, filename: str) -> List[dict]:
        """
        Parse an Excel file containing grants.
        Returns list of grant dictionaries.
        """
        try:
            workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            sheet = workbook.active

            if sheet is None:
                raise ValueError("No active sheet found in Excel file")

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).lower().strip() if cell.value else "")

            # Map headers to our fields
            field_mapping = {}
            for our_field, possible_names in GrantService.COLUMN_MAPPINGS.items():
                for i, header in enumerate(headers):
                    if header in possible_names:
                        field_mapping[our_field] = i
                        break

            # Parse grants
            grants = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                grant = {
                    "row_number": row_idx,
                    "raw_data": {headers[i]: str(v) if v else None for i, v in enumerate(row) if i < len(headers)}
                }

                # Map known fields
                for our_field, col_idx in field_mapping.items():
                    if col_idx < len(row):
                        value = row[col_idx]
                        grant[our_field] = GrantService._process_field(our_field, value)

                # Only add if has a name
                if grant.get("name"):
                    grants.append(grant)

            return grants

        except Exception as e:
            raise ValueError(f"Failed to parse Excel file: {str(e)}")

    @staticmethod
    def _process_field(field_name: str, value) -> Optional[any]:
        """Process a field value based on its type."""
        if value is None:
            return None

        if field_name in ["amount_min", "amount_max"]:
            return GrantService._parse_amount(value)
        elif field_name == "deadline":
            return GrantService._parse_deadline(value)
        elif field_name == "categories":
            return GrantService._parse_categories(value)
        elif field_name == "eligibility":
            return GrantService._parse_eligibility(value)
        else:
            return str(value).strip() if value else None

    @staticmethod
    def _parse_amount(value) -> Optional[float]:
        """Parse an amount field."""
        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            # Remove common formatting
            cleaned = value.replace("$", "").replace(",", "").strip()

            # Handle "up to X" format
            if "up to" in cleaned.lower():
                cleaned = cleaned.lower().replace("up to", "").strip()

            try:
                return float(cleaned)
            except ValueError:
                pass

        return None

    @staticmethod
    def _parse_deadline(value) -> Optional[str]:
        """Parse a deadline field."""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

        if isinstance(value, str):
            value_lower = value.lower().strip()

            # Check for rolling deadline
            if "rolling" in value_lower or "ongoing" in value_lower or "open" in value_lower:
                return "rolling"

            # Try to parse date
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y"]:
                try:
                    parsed = datetime.strptime(value.strip(), fmt)
                    return parsed.strftime("%Y-%m-%d")
                except ValueError:
                    continue

            return value.strip()

        return None

    @staticmethod
    def _parse_categories(value) -> Optional[List[str]]:
        """Parse categories field."""
        if isinstance(value, str):
            # Split by comma, semicolon, or pipe
            import re
            parts = re.split(r'[,;|]', value)
            return [p.strip() for p in parts if p.strip()]
        return None

    @staticmethod
    def _parse_eligibility(value) -> Optional[dict]:
        """Parse eligibility field into structured format."""
        if isinstance(value, str):
            text = value.lower()
            eligibility = {
                "raw_text": value,
                "requires_501c3": "501(c)(3)" in value or "501c3" in text,
                "requires_catholic": "catholic" in text,
                "requires_school": "school" in text and "must" in text,
            }
            return eligibility
        return None

    @staticmethod
    def format_amount_display(amount_min: Optional[float], amount_max: Optional[float]) -> str:
        """Format amount for display."""
        def fmt(val):
            if val >= 1000000:
                return f"${val/1000000:.1f}M"
            elif val >= 1000:
                return f"${val/1000:.0f}K"
            else:
                return f"${val:,.0f}"

        if amount_min and amount_max:
            if amount_min == amount_max:
                return fmt(amount_min)
            return f"{fmt(amount_min)} - {fmt(amount_max)}"
        elif amount_max:
            return f"Up to {fmt(amount_max)}"
        elif amount_min:
            return f"From {fmt(amount_min)}"
        return "Varies"

    @staticmethod
    def format_deadline_display(deadline: Optional[str], deadline_type: Optional[str] = None) -> tuple[str, bool]:
        """
        Format deadline for display.
        Returns (display_string, is_urgent)
        """
        if not deadline:
            return "TBD — check website", False

        if deadline.lower() == "rolling":
            return "Rolling deadline", False

        try:
            deadline_date = datetime.strptime(deadline, "%Y-%m-%d")
            now = datetime.now()
            days_until = (deadline_date - now).days

            if days_until < 0:
                return f"Closed: {deadline_date.strftime('%b %d, %Y')}", False
            elif days_until <= 14:
                return f"Due: {deadline_date.strftime('%b %d, %Y')} ({days_until} days)", True
            elif days_until <= 30:
                return f"Due: {deadline_date.strftime('%b %d, %Y')} ({days_until} days)", True
            else:
                return f"Due: {deadline_date.strftime('%b %d, %Y')}", False

        except ValueError:
            return deadline, False

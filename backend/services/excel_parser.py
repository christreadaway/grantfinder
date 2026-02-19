"""
Excel parser for grant database.
Handles 5-category structure per v2.6 spec.
"""
import io
import uuid
from datetime import datetime
from typing import Dict, List, Any
import logging

from openpyxl import load_workbook

from models.schemas import (
    Grant, GrantCategory, GrantStatus, GeoQualified,
    Foundation
)

logger = logging.getLogger(__name__)

# Expected sheet names and their categories
CATEGORY_SHEETS = {
    "Church/Parish Grants": GrantCategory.CHURCH_PARISH,
    "Category 1": GrantCategory.CHURCH_PARISH,
    "Parish Grants": GrantCategory.CHURCH_PARISH,

    "Catholic School Grants": GrantCategory.CATHOLIC_SCHOOL,
    "Category 2": GrantCategory.CATHOLIC_SCHOOL,
    "School Grants": GrantCategory.CATHOLIC_SCHOOL,

    "Mixed Church-School": GrantCategory.MIXED_CHURCH_SCHOOL,
    "Category 3": GrantCategory.MIXED_CHURCH_SCHOOL,
    "Mixed": GrantCategory.MIXED_CHURCH_SCHOOL,

    "Non-Catholic Qualifying": GrantCategory.NON_CATHOLIC_QUALIFYING,
    "Category 4": GrantCategory.NON_CATHOLIC_QUALIFYING,
    "Non-Catholic": GrantCategory.NON_CATHOLIC_QUALIFYING,
    "Secular Grants": GrantCategory.NON_CATHOLIC_QUALIFYING,

    "Catholic Foundations": GrantCategory.CATHOLIC_FOUNDATIONS,
    "Category 5": GrantCategory.CATHOLIC_FOUNDATIONS,
    "Foundations": GrantCategory.CATHOLIC_FOUNDATIONS,
}

# Required columns for grants (v2.6)
REQUIRED_GRANT_COLUMNS = [
    "grant_name", "deadline", "amount", "funder", "description",
    "contact", "url", "status", "geo_qualified"
]

# Required columns for foundations
REQUIRED_FOUNDATION_COLUMNS = [
    "foundation_name", "application_cycle", "focus_areas",
    "location", "contact", "website", "annual_giving"
]

# Column name mapping (flexible matching)
COLUMN_MAP = {
    # Grant columns
    "grant name": "grant_name",
    "grant_name": "grant_name",
    "name": "grant_name",

    "deadline": "deadline",
    "due date": "deadline",
    "due_date": "deadline",

    "amount": "amount",
    "grant amount": "amount",
    "award": "amount",

    "funder": "funder",
    "funding org": "funder",
    "organization": "funder",

    "description": "description",
    "desc": "description",
    "details": "description",

    "contact": "contact",
    "contact info": "contact",
    "email": "contact",

    "url": "url",
    "link": "url",
    "website": "url",
    "application url": "url",

    "status": "status",
    "grant status": "status",

    "geo qualified": "geo_qualified",
    "geo_qualified": "geo_qualified",
    "geographic": "geo_qualified",
    "geography": "geo_qualified",

    "funder stats": "funder_stats",
    "funder_stats": "funder_stats",
    "stats": "funder_stats",

    # Foundation columns
    "foundation name": "foundation_name",
    "foundation_name": "foundation_name",

    "application cycle": "application_cycle",
    "application_cycle": "application_cycle",
    "cycle": "application_cycle",

    "focus areas": "focus_areas",
    "focus_areas": "focus_areas",
    "focus": "focus_areas",

    "location": "location",
    "city": "location",

    "annual giving": "annual_giving",
    "annual_giving": "annual_giving",
    "giving": "annual_giving",

    "notes": "notes",
}


def normalize_column_name(name: str) -> str:
    """Normalize column name for matching."""
    if name is None:
        return ""
    normalized = name.lower().strip()
    return COLUMN_MAP.get(normalized, normalized.replace(" ", "_"))


def parse_status(value: str) -> GrantStatus:
    """Parse grant status string."""
    if value is None:
        return GrantStatus.CHECK_DEADLINE

    value = str(value).upper().strip()

    if "OPEN" in value:
        return GrantStatus.OPEN
    elif "ROLL" in value:
        return GrantStatus.ROLLING
    elif "CLOSE" in value:
        return GrantStatus.CLOSED
    else:
        return GrantStatus.CHECK_DEADLINE


def parse_geo_qualified(value: str) -> GeoQualified:
    """Parse geographic qualification string."""
    if value is None:
        return GeoQualified.CHECK

    value = str(value).upper().strip()

    if "TX" in value or "TEXAS" in value:
        return GeoQualified.TX_ONLY
    elif value in ("YES", "Y", "TRUE", "1"):
        return GeoQualified.YES
    elif value in ("NO", "N", "FALSE", "0"):
        return GeoQualified.NO
    else:
        return GeoQualified.CHECK


async def parse_grant_database(
    file_content: bytes,
    user_id: str
) -> Dict[str, Any]:
    """
    Parse Excel grant database with 5 categories.

    Returns:
        Dict with grants, foundations, category_counts, and upload_id
    """
    upload_id = str(uuid.uuid4())
    grants: List[Grant] = []
    foundations: List[Foundation] = []
    category_counts: Dict[str, int] = {cat.value: 0 for cat in GrantCategory}

    try:
        # Load workbook
        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        logger.info(f"Workbook sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            # Determine category from sheet name
            category = None
            for pattern, cat in CATEGORY_SHEETS.items():
                if pattern.lower() in sheet_name.lower():
                    category = cat
                    break

            if category is None:
                logger.warning(f"Unknown sheet category: {sheet_name}, skipping")
                continue

            sheet = wb[sheet_name]

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(normalize_column_name(cell.value))

            logger.info(f"Sheet '{sheet_name}' headers: {headers}")

            # Parse rows
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(row):
                    continue

                # Create row dict
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_data[headers[idx]] = value

                # Handle foundations (Category 5) differently
                if category == GrantCategory.CATHOLIC_FOUNDATIONS:
                    foundation = parse_foundation_row(row_data, user_id, upload_id)
                    if foundation:
                        foundations.append(foundation)
                else:
                    grant = parse_grant_row(row_data, user_id, category, upload_id)
                    if grant:
                        grants.append(grant)
                        category_counts[category.value] += 1

        logger.info(
            f"Parsed {len(grants)} grants, {len(foundations)} foundations "
            f"from {len(wb.sheetnames)} sheets"
        )

        return {
            "grants": grants,
            "foundations": foundations,
            "category_counts": category_counts,
            "upload_id": upload_id,
        }

    except Exception as e:
        logger.error(f"Excel parsing error: {e}")
        raise ValueError(f"Failed to parse Excel file: {str(e)}")


def parse_grant_row(
    row_data: Dict[str, Any],
    user_id: str,
    category: GrantCategory,
    upload_id: str
) -> Grant:
    """Parse a single grant row."""
    try:
        # Extract required fields with defaults
        grant_name = str(row_data.get("grant_name", "") or "").strip()
        if not grant_name:
            return None

        grant = Grant(
            id=f"grant_{uuid.uuid4().hex[:12]}",
            user_id=user_id,
            grant_name=grant_name,
            deadline=str(row_data.get("deadline", "Check website") or "Check website"),
            amount=str(row_data.get("amount", "Varies") or "Varies"),
            funder=str(row_data.get("funder", "Unknown") or "Unknown"),
            description=str(row_data.get("description", "") or ""),
            contact=str(row_data.get("contact", "") or "See website"),
            url=str(row_data.get("url", "") or ""),
            status=parse_status(row_data.get("status")),
            geo_qualified=parse_geo_qualified(row_data.get("geo_qualified")),
            funder_stats=str(row_data.get("funder_stats", "") or "") if row_data.get("funder_stats") else None,
            category=category,
            created_at=datetime.utcnow(),
        )

        return grant

    except Exception as e:
        logger.warning(f"Failed to parse grant row: {e}")
        return None


def parse_foundation_row(
    row_data: Dict[str, Any],
    user_id: str,
    upload_id: str
) -> Foundation:
    """Parse a foundation row (Category 5)."""
    try:
        foundation_name = str(row_data.get("foundation_name", "") or "").strip()
        if not foundation_name:
            return None

        foundation = Foundation(
            id=f"foundation_{uuid.uuid4().hex[:12]}",
            user_id=user_id,
            foundation_name=foundation_name,
            application_cycle=str(row_data.get("application_cycle", "Check website") or "Check website"),
            focus_areas=str(row_data.get("focus_areas", "") or ""),
            location=str(row_data.get("location", "") or ""),
            contact=str(row_data.get("contact", "") or "See website"),
            website=str(row_data.get("website", "") or row_data.get("url", "") or ""),
            annual_giving=str(row_data.get("annual_giving", "") or ""),
            notes=str(row_data.get("notes", "") or "") if row_data.get("notes") else None,
            created_at=datetime.utcnow(),
        )

        return foundation

    except Exception as e:
        logger.warning(f"Failed to parse foundation row: {e}")
        return None
